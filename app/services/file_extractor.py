"""
app/services/file_extractor.py
───────────────────────────────
Text extraction from uploaded vendor documents.

Supports: PDF, DOCX, TXT, XLSX, CSV, HTML
Handles: plain paths, base64-encoded paths, URL-encoded paths.

Single Responsibility: extracts text from a file given its path.
Open/Closed: new file types added by registering an extractor function.

Fallback chain for path resolution:
  1. Try path as-is (plain string)
  2. Try base64 decoding
  3. Try URL decoding
  4. Log error and return empty
"""

from __future__ import annotations

import base64
import csv
import io
import logging
import os
from pathlib import Path
from typing import Callable, Dict, Optional
from urllib.parse import unquote

log = logging.getLogger("app.services.file_extractor")

_EXTRACTORS: Dict[str, Callable[[str], str]] = {}


def _register(extensions: list[str]):
    """Decorator to register an extractor for given file extensions."""
    def decorator(fn: Callable[[str], str]):
        for ext in extensions:
            _EXTRACTORS[ext.lower().lstrip(".")] = fn
        return fn
    return decorator


# ── Path resolution ─────────────────────────────────────────────

def resolve_file_path(raw_path: str) -> Optional[str]:
    """
    Attempt to resolve a file path that may be plain, base64-encoded,
    or URL-encoded. Returns the resolved path if the file exists, else None.
    """
    if not raw_path or not raw_path.strip():
        return None

    cleaned = raw_path.strip()

    if os.path.isfile(cleaned):
        return cleaned

    try:
        decoded_b64 = base64.b64decode(cleaned, validate=True).decode("utf-8", errors="replace")
        if os.path.isfile(decoded_b64):
            log.debug("Resolved base64-encoded path: %s", decoded_b64)
            return decoded_b64
    except Exception:
        pass

    try:
        decoded_url = unquote(cleaned)
        if decoded_url != cleaned and os.path.isfile(decoded_url):
            log.debug("Resolved URL-encoded path: %s", decoded_url)
            return decoded_url
    except Exception:
        pass

    common_prefixes = ["/var/www/", "/opt/ofb/uploads/", "/home/", "./uploads/", "../uploads/"]
    basename = Path(cleaned).name
    for prefix in common_prefixes:
        candidate = os.path.join(prefix, basename)
        if os.path.isfile(candidate):
            log.debug("Resolved via prefix scan: %s", candidate)
            return candidate

    return None


def get_file_extension(file_path: str, file_type: str = "") -> str:
    """Determine file extension from path or explicit type field."""
    if file_type and file_type.strip():
        return file_type.strip().lower().lstrip(".")

    _, ext = os.path.splitext(file_path)
    return ext.lower().lstrip(".")


# ── Extractors ──────────────────────────────────────────────────

@_register(["txt", "text", "log", "md", "rst"])
def _extract_text(file_path: str) -> str:
    encodings = ["utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    return ""


@_register(["csv", "tsv"])
def _extract_csv(file_path: str) -> str:
    lines = []
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                lines.append(" | ".join(row))
    except Exception as exc:
        log.warning("CSV extraction failed for %s: %s", file_path, exc)
    return "\n".join(lines)


@_register(["pdf"])
def _extract_pdf(file_path: str) -> str:
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        log.warning("PyPDF2 not installed — cannot extract PDF: %s", file_path)
        return ""

    try:
        reader = PdfReader(file_path)
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages)
    except Exception as exc:
        log.warning("PDF extraction failed for %s: %s", file_path, exc)
        return ""


@_register(["docx"])
def _extract_docx(file_path: str) -> str:
    try:
        from docx import Document
    except ImportError:
        log.warning("python-docx not installed — cannot extract DOCX: %s", file_path)
        return ""

    try:
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)
    except Exception as exc:
        log.warning("DOCX extraction failed for %s: %s", file_path, exc)
        return ""


@_register(["xlsx", "xls"])
def _extract_excel(file_path: str) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("openpyxl not installed — cannot extract Excel: %s", file_path)
        return ""

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                cell_texts = [str(c) if c is not None else "" for c in row]
                if any(cell_texts):
                    lines.append(" | ".join(cell_texts))
        wb.close()
        return "\n".join(lines)
    except Exception as exc:
        log.warning("Excel extraction failed for %s: %s", file_path, exc)
        return ""


@_register(["html", "htm"])
def _extract_html(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except Exception as exc:
        log.warning("HTML read failed for %s: %s", file_path, exc)
        return ""

    text_parts = []
    in_tag = False
    current = []
    for char in content:
        if char == "<":
            if current:
                text_parts.append("".join(current))
                current = []
            in_tag = True
        elif char == ">":
            in_tag = False
        elif not in_tag:
            current.append(char)
    if current:
        text_parts.append("".join(current))

    return " ".join(text_parts).strip()


# ── Public API ──────────────────────────────────────────────────

def extract_text_from_file(
    file_path: str,
    file_type: str = "",
) -> Optional[str]:
    """
    Extract text from a file. Returns the extracted text, or None on failure.

    Handles path resolution (plain, base64, URL-encoded) and dispatches
    to the appropriate extractor based on file extension.
    """
    resolved = resolve_file_path(file_path)
    if resolved is None:
        log.debug("Could not resolve file path: %s", file_path[:100])
        return None

    ext = get_file_extension(resolved, file_type)
    extractor = _EXTRACTORS.get(ext)

    if extractor is None:
        log.debug("No extractor registered for extension '%s': %s", ext, resolved)
        return None

    try:
        text = extractor(resolved)
        if text and text.strip():
            return text.strip()
        return None
    except Exception as exc:
        log.error("Extraction failed for %s: %s", resolved, exc)
        return None


def supported_extensions() -> list[str]:
    """Return list of supported file extensions."""
    return sorted(_EXTRACTORS.keys())
